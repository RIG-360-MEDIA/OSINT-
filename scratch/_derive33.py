import collections
import csv
import openpyxl

P = r"D:\global_news_dataset_PERFECTED_20260527 (2).xlsx"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
mrows = list(wb["00 - MASTER INDEX"].iter_rows(values_only=True))
mi = {h: i for i, h in enumerate(mrows[0])}

small = [r for r in mrows[1:] if "LOW_COVERAGE" in str(r[mi["Notes"]] or "").upper()]
print("MASTER countries:", len(mrows) - 1)
print("LOW_COVERAGE_REVIEW states:", len(small))
print("sum master Live_Sites over LOW_COVERAGE:", sum(int(r[mi["Live_Sites"]] or 0) for r in small))
print("\nISO2 | Country | Live | GeoBlk | Cov | Notes")
for r in sorted(small, key=lambda x: int(x[mi["Live_Sites"]] or 0)):
    print(r[mi["ISO_A2"]], "|", str(r[mi["Country"]])[:26], "|", r[mi["Live_Sites"]], "|",
          r[mi["Geo_Blocked_Sites"]], "|", r[mi["Coverage_Score"]], "|", r[mi["Notes"]])

iso_set = {r[mi["ISO_A2"]] for r in small}
sheet_for = {s.split(" - ")[0].strip(): s for s in wb.sheetnames if " - " in s}

out, percountry, allblocked = [], {}, []
for iso in sorted(iso_set):
    sh = sheet_for.get(iso)
    if not sh:
        print("NO SHEET for", iso); continue
    rows = list(wb[sh].iter_rows(values_only=True))
    h = {str(c).strip(): i for i, c in enumerate(rows[0])}
    live = 0
    for r in rows[1:]:
        if str(r[h["access_status"]]).strip() == "LIVE":
            live += 1
            out.append({"iso2": iso, "name": r[h["website_name"]], "url": r[h["url"]],
                        "final_url": r[h["final_url"]], "language": r[h["language"]],
                        "category": r[h["category"]], "reach_tier": r[h["reach_tier"]]})
    percountry[iso] = (live, len(rows) - 1)
    if len(rows) - 1 > 0 and live == 0:
        allblocked.append(iso)

print("\nTOTAL LIVE candidate sources (LOW_COVERAGE set):", len(out))
print("ALL-BLOCKED states (rows exist, 0 LIVE):", allblocked or "none")
print("per-country (iso: live/total):", {k: f"{v[0]}/{v[1]}" for k, v in sorted(percountry.items())})
with open(r"C:\Users\Dell\Desktop\rig-surveillance\scratch\_candidate_sources.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["iso2", "name", "url", "final_url", "language", "category", "reach_tier"])
    w.writeheader(); w.writerows(out)
print("wrote scratch/_candidate_sources.csv (", len(out), "rows )")
# distinct languages (some non-English per the spec)
print("languages:", dict(collections.Counter(o["language"] for o in out)))
