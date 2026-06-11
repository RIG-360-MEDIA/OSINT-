import collections
import csv
import openpyxl
from urllib.parse import urlparse

ROOT = r"C:\Users\Dell\Desktop\rig-surveillance\scratch"
P = r"D:\global_news_dataset_PERFECTED_20260527 (2).xlsx"


def dom(u):
    if not u:
        return ""
    try:
        n = urlparse(u if "://" in str(u) else "http://" + str(u)).netloc.lower()
    except Exception:
        n = ""
    return n[4:] if n.startswith("www.") else n


# our coverage
our_count = collections.Counter()
our_domains = set()
with open(ROOT + r"\_our_sources.csv", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line:
            continue
        p = line.split(",", 1)
        c = p[0].strip()
        d = (p[1].strip().lower() if len(p) > 1 else "")
        if c:
            our_count[c] += 1
        if d:
            our_domains.add(d[4:] if d.startswith("www.") else d)

# file LIVE per country
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
mrows = list(wb["00 - MASTER INDEX"].iter_rows(values_only=True))
mi = {h: i for i, h in enumerate(mrows[0])}
cname = {r[mi["ISO_A2"]]: r[mi["Country"]] for r in mrows[1:]}
sheet_for = {s.split(" - ")[0].strip(): s for s in wb.sheetnames if " - " in s}

file_live = collections.Counter()
live_rows = collections.defaultdict(list)
for iso, sh in sheet_for.items():
    rows = list(wb[sh].iter_rows(values_only=True))
    if not rows:
        continue
    h = {str(c).strip(): i for i, c in enumerate(rows[0])}
    if "access_status" not in h:
        continue
    for r in rows[1:]:
        if str(r[h["access_status"]]).strip() == "LIVE":
            file_live[iso] += 1
            live_rows[iso].append({"name": r[h["website_name"]], "url": r[h["url"]],
                                   "language": r[h["language"]], "category": r[h["category"]],
                                   "reach_tier": r[h["reach_tier"]]})

cut0 = sorted(i for i in file_live if our_count[i] == 0)
cut1 = sorted(i for i in file_live if our_count[i] <= 1)
print("file countries with LIVE>0:", len(file_live))
print("CUT our==0 & fileLIVE>0:", len(cut0), "| sum fileLIVE:", sum(file_live[i] for i in cut0))
print("CUT our<=1 & fileLIVE>0:", len(cut1), "| sum fileLIVE:", sum(file_live[i] for i in cut1),
      "| our sources in this set:", sum(our_count[i] for i in cut1))
print("\niso our file_live country   [CUT our==0]")
for i in sorted(cut0, key=lambda x: -file_live[x]):
    print(f"  {i}  {our_count[i]}  {file_live[i]:>2}  {str(cname.get(i))[:24]}")

# dedup candidate import set for cut1 (broader), by domain
seen = set()
imp = []
dupes = 0
for iso in cut1:
    for s in live_rows[iso]:
        d = dom(s["url"])
        if d and d in our_domains:
            dupes += 1
            continue
        if d in seen:
            continue
        seen.add(d)
        imp.append({"iso2": iso, "name": s["name"], "domain": d, "url": s["url"],
                    "language": s["language"], "category": s["category"], "reach_tier": s["reach_tier"]})
with open(ROOT + r"\_import_set.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["iso2", "name", "domain", "url", "language", "category", "reach_tier"])
    w.writeheader(); w.writerows(imp)
print(f"\nimport-set (cut<=1, deduped): {len(imp)} rows  | domain-dupes skipped: {dupes}")
print("languages in import-set:", dict(collections.Counter(r["language"] for r in imp)))
print("wrote scratch/_import_set.csv")
