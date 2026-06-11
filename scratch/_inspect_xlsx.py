import collections
import openpyxl

P = r"D:\global_news_dataset_PERFECTED_20260527 (2).xlsx"
wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
sn = wb.sheetnames
print("N_SHEETS:", len(sn))
print("SHEETS:", sn)

# MASTER INDEX sheet
master = next((s for s in sn if "MASTER" in s.upper() or "INDEX" in s.upper()), sn[0])
print("\n=== MASTER:", master, "===")
mrows = list(wb[master].iter_rows(values_only=True))
print("HEADERS:", mrows[0])
print("NROWS:", len(mrows))
for r in mrows[1:9]:
    print("  ", r)

# a per-country sheet (first non-master)
csheet = next((s for s in sn if s != master), None)
if csheet:
    crows = list(wb[csheet].iter_rows(values_only=True))
    print("\n=== COUNTRY SHEET:", csheet, "===")
    print("HEADERS:", crows[0])
    print("NROWS:", len(crows))
    for r in crows[1:4]:
        print("  ", r)

# access_status distribution across ALL per-country sheets (find the column by header)
acc = collections.Counter()
percountry_total = 0
for s in sn:
    if s == master:
        continue
    rows = list(wb[s].iter_rows(values_only=True))
    if not rows:
        continue
    hdr = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    ai = hdr.index("access_status") if "access_status" in hdr else None
    for r in rows[1:]:
        percountry_total += 1
        if ai is not None and ai < len(r):
            acc[str(r[ai]).strip()] += 1
print("\n=== ACROSS ALL PER-COUNTRY SHEETS ===")
print("total data rows:", percountry_total)
print("access_status counts:", dict(acc))
