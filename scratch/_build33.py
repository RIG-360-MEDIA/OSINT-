import collections
import csv
import openpyxl
from urllib.parse import urlparse

ROOT = r"C:\Users\Dell\Desktop\rig-surveillance\scratch"
P = r"D:\global_news_dataset_PERFECTED_20260527 (2).xlsx"
ISO33 = ["AG", "BS", "BB", "BZ", "BW", "BN", "CY", "DM", "SZ", "FJ", "GA", "GM", "GD", "GY",
         "JM", "KI", "LS", "MV", "MT", "MU", "NA", "NR", "KN", "LC", "VC", "WS", "SC", "SB",
         "TO", "TT", "TV", "VU", "ZM"]
assert len(ISO33) == 33, len(ISO33)


def dom(u):
    if not u:
        return ""
    n = urlparse(str(u) if "://" in str(u) else "http://" + str(u)).netloc.lower()
    return n[4:] if n.startswith("www.") else n


our_domains, our_count = set(), collections.Counter()
with open(ROOT + r"\_our_sources.csv", encoding="utf-8") as f:
    for line in f:
        p = line.strip().split(",", 1)
        if p[0].strip():
            our_count[p[0].strip()] += 1
        if len(p) > 1 and p[1].strip():
            d = p[1].strip().lower()
            our_domains.add(d[4:] if d.startswith("www.") else d)

wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
sheet_for = {s.split(" - ")[0].strip(): s for s in wb.sheetnames if " - " in s}

rows_out, per, allblocked, missing, langs = [], {}, [], [], collections.Counter()
seen = set()
for iso in ISO33:
    sh = sheet_for.get(iso)
    if not sh:
        missing.append(iso); continue
    rs = list(wb[sh].iter_rows(values_only=True))
    h = {str(c).strip(): i for i, c in enumerate(rs[0])}
    live = added = dup = 0
    for r in rs[1:]:
        if str(r[h["access_status"]]).strip() != "LIVE":
            continue
        live += 1
        d = dom(r[h["url"]])
        if not d or d in our_domains or d in seen:
            dup += 1; continue
        seen.add(d)
        added += 1
        lang = (str(r[h["language"]]).strip().lower() if r[h["language"]] else "en")[:8]
        langs[lang] += 1
        cat = (str(r[h["category"]]).strip() if r[h["category"]] else "general")
        tier = r[h["reach_tier"]] if isinstance(r[h["reach_tier"]], int) else 3
        rows_out.append({"name": r[h["website_name"]], "domain": d, "rss_url": r[h["url"]],
                         "source_type": "pending_feed", "source_tier": tier, "language": lang,
                         "country": iso, "topics": "{" + cat + "}", "is_active": "false"})
    per[iso] = {"file_live": live, "added": added, "dup": dup, "ours_now": our_count[iso]}
    if live == 0 and (len(rs) - 1) > 0:
        allblocked.append(iso)

with open(ROOT + r"\_import33.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["name", "domain", "rss_url", "source_type", "source_tier",
                                      "language", "country", "topics", "is_active"])
    w.writeheader(); w.writerows(rows_out)

print("missing sheets:", missing or "none")
print("all-blocked (0 LIVE) within 33:", allblocked or "none")
print("TOTAL file-LIVE:", sum(p["file_live"] for p in per.values()),
      "| TOTAL to add (deduped):", len(rows_out),
      "| dups skipped:", sum(p["dup"] for p in per.values()))
print("languages:", dict(langs))
print("\niso file_live add dup ours_now")
for iso in ISO33:
    p = per.get(iso, {})
    print(f"  {iso}  {p.get('file_live','-'):>2}  {p.get('added','-'):>2}  {p.get('dup','-'):>2}  {p.get('ours_now',0)}")
print("wrote scratch/_import33.csv")
